"""Layer D — selection + sheet assembly (enrichment_validation_dataset_plan.md §6)."""
from __future__ import annotations

import json

import openpyxl
import pytest

from app.dataset import (
    MAX_PER_CLASS,
    ProductionCandidate,
    _discovery_class_info,
    _score_claims,
    _urgency_tier_rank,
    build_candidate,
    gather_survivors,
    persist_selection,
    select_50,
    write_workbook,
)
from app.db import (
    add_entity_source,
    connection,
    get_production_records,
    upsert_claims,
    upsert_entity,
    write_audit_rejected_value,
    write_rejection,
)


def _claim(field_name, status="confirmed", confidence="medium", answer="x", **kw):
    d = {"field_name": field_name, "answer": answer, "status": status, "confidence": confidence}
    d.update(kw)
    return d


# --- scoring ---

def test_score_claims_weighs_verified_above_single_source():
    verified_score, verified_count = _score_claims([_claim("f1", status="verified")])
    single_score, single_count = _score_claims([_claim("f1", status="single_source")])
    assert verified_score > single_score
    assert verified_count == 1
    assert single_count == 0


def test_score_claims_ignores_claims_without_field_name():
    score, _ = _score_claims([{"question_id": "G1.Q1", "answer": "x", "status": "confirmed", "confidence": "high"}])
    assert score == 0


def test_urgency_tier_rank_levels():
    assert _urgency_tier_rank([]) == 0
    assert _urgency_tier_rank([_claim("important_insight", confidence="low")]) == 1
    assert _urgency_tier_rank([_claim("important_insight", confidence="high")]) == 2


def test_discovery_class_info_picks_first_alphabetically():
    claims = [_claim("discovery_class_fec_employer"), _claim("discovery_class_13f_filing")]
    primary, count = _discovery_class_info(claims)
    assert primary == "13f_filing"
    assert count == 2


# --- select_50 quota logic ---

def _candidate(entity_id, cls, score=1.0, type_final="SFO"):
    return ProductionCandidate(
        entity_id=entity_id, canonical_name=entity_id, type_final=type_final, outcome="ship",
        claims=[], actionability_score=score, verified_cell_count=0, urgency_tier_rank=0,
        discovery_class_primary=cls, discovery_class_count=1,
    )


def test_select_50_ranks_by_actionability_score_descending():
    candidates = [_candidate("low", "a", score=1.0), _candidate("high", "b", score=10.0)]
    selected, _, _ = select_50(candidates, n=2)
    assert [c.entity_id for c in selected] == ["high", "low"]


def test_select_50_caps_at_max_per_class():
    candidates = [_candidate(f"e{i}", "same_class", score=float(60 - i)) for i in range(60)]
    selected, per_class, excluded = select_50(candidates, n=50)
    assert per_class["same_class"] == MAX_PER_CLASS
    assert len(selected) == MAX_PER_CLASS  # only one class exists -> quota is the real cap
    assert len(excluded) == 60 - MAX_PER_CLASS


def test_select_50_diversifies_across_classes():
    candidates = (
        [_candidate(f"a{i}", "class_a", score=float(100 - i)) for i in range(20)]
        + [_candidate(f"b{i}", "class_b", score=float(50 - i)) for i in range(20)]
    )
    selected, per_class, _ = select_50(candidates, n=30)
    assert per_class["class_a"] <= MAX_PER_CLASS
    assert per_class["class_b"] <= MAX_PER_CLASS
    assert len(selected) == 30


def test_select_50_prefers_sfo_over_mfo_on_tied_score():
    candidates = [_candidate("mfo_lead", "a", score=5.0, type_final="MFO"),
                  _candidate("sfo_lead", "b", score=5.0, type_final="SFO")]
    selected, _, _ = select_50(candidates, n=2)
    assert selected[0].entity_id == "sfo_lead"


# --- build_candidate + gather_survivors (DB-integrated) ---

def test_build_candidate_from_real_claims(db_path):
    with connection(db_path) as conn:
        upsert_entity(conn, "e1", "Acme Capital")
        upsert_claims(conn, "e1", [
            _claim("principal_name", status="verified", answer="Jane Doe"),
            _claim("aum_usd", status="single_source", answer=100_000_000),
            _claim("discovery_class_13f_filing", answer=True, status="confirmed"),
        ])
        entity = {"entity_id": "e1", "canonical_name": "Acme Capital"}
        from app.db import get_claims
        claims = get_claims(conn, "e1")
    candidate = build_candidate(entity, claims, outcome="ship", type_final="SFO")
    assert candidate.verified_cell_count == 1
    assert candidate.discovery_class_primary == "13f_filing"
    assert candidate.actionability_score > 0


def test_gather_survivors_skips_unknown_entities(db_path):
    with connection(db_path) as conn:
        upsert_entity(conn, "e1", "Acme Capital")
        upsert_claims(conn, "e1", [_claim("principal_name", answer="Jane Doe")])
        candidates = gather_survivors(conn, [("e1", "ship", "SFO"), ("nonexistent", "ship", "SFO")])
    assert len(candidates) == 1
    assert candidates[0].entity_id == "e1"


def test_persist_selection_writes_production_records(db_path):
    with connection(db_path) as conn:
        upsert_entity(conn, "e1", "Selected Co")
        upsert_entity(conn, "e2", "Excluded Co")
        selected = [_candidate("e1", "class_a")]
        excluded = [_candidate("e2", "class_a")]
        persist_selection(conn, selected, excluded)
    with connection(db_path) as conn:
        records = get_production_records(conn)
    by_id = {r["entity_id"]: r for r in records}
    assert by_id["e1"]["excluded_by_quota"] is False
    assert by_id["e1"]["rank"] == 1
    assert by_id["e2"]["excluded_by_quota"] is True


# --- write_workbook: six sheets + CSV + manifest ---

def test_write_workbook_produces_six_sheets_and_csv_and_manifest(db_path, tmp_path):
    with connection(db_path) as conn:
        upsert_entity(conn, "e1", "Acme Capital")
        upsert_claims(conn, "e1", [
            _claim("principal_name", status="verified", answer="Jane Doe", source_url="http://a"),
            _claim("principal_email", status="could_not_verify", answer=None),
            _claim("aum_usd", status="single_source", answer=100_000_000, source_class="13f_filing"),
            _claim("discovery_class_13f_filing", answer=True, status="confirmed"),
        ])
        write_audit_rejected_value(conn, "e1", "principal_phone", "555-0000", "verification_failed")
        upsert_entity(conn, "e2", "Rejected Co")
        write_rejection(conn, "e2", reason_code="V6_completeness", gate_results={}, claim_ledger=[], stage="validation")

        from app.db import get_claims as _get_claims
        entity = {"entity_id": "e1", "canonical_name": "Acme Capital"}
        candidate = build_candidate(entity, _get_claims(conn, "e1"), outcome="ship", type_final="SFO")
        selected, per_class, excluded = select_50([candidate], n=50)
        persist_selection(conn, selected, excluded)

        paths = write_workbook(conn, selected, per_class, excluded, tmp_path,
                                budget_spent={"usd_spent": 1.23})

    wb = openpyxl.load_workbook(paths["xlsx"])
    assert set(wb.sheetnames) == {
        "records", "provenance", "audit_rejected_values", "rejected_records",
        "source_class_report", "data_dictionary",
    }

    records_ws = wb["records"]
    header = [c.value for c in next(records_ws.iter_rows(max_row=1))]
    assert "principal_name" in header
    assert "principal_email_status" in header  # high-value field companion status column
    row = [c.value for c in next(records_ws.iter_rows(min_row=2, max_row=2))]
    row_dict = dict(zip(header, row))
    assert row_dict["principal_name"] == "Jane Doe"
    assert row_dict["principal_email"] is None
    assert row_dict["principal_email_status"] == "could_not_verify"  # never silently empty

    audit_ws = wb["audit_rejected_values"]
    audit_rows = list(audit_ws.iter_rows(min_row=2, values_only=True))
    assert any(r[1] == "principal_phone" for r in audit_rows)

    rejected_ws = wb["rejected_records"]
    rejected_rows = list(rejected_ws.iter_rows(min_row=2, values_only=True))
    assert any(r[0] == "e2" and r[1] == "validation" for r in rejected_rows)

    import csv as _csv
    with open(paths["csv"], newline="", encoding="utf-8") as f:
        csv_rows = list(_csv.DictReader(f))
    assert csv_rows[0]["principal_name"] == "Jane Doe"

    manifest = json.loads(open(paths["manifest"], encoding="utf-8").read())
    assert manifest["selected_count"] == 1
    assert manifest["budget_spent"]["usd_spent"] == 1.23
    assert "git_sha" in manifest


def test_write_workbook_excluded_by_quota_appears_in_rejected_records(db_path, tmp_path):
    candidates = [_candidate(f"e{i}", "class_a", score=float(20 - i)) for i in range(18)]
    for c in candidates:
        with connection(db_path) as conn:
            upsert_entity(conn, c.entity_id, c.entity_id)
    selected, per_class, excluded = select_50(candidates, n=50, max_per_class=15)
    assert len(excluded) == 3
    with connection(db_path) as conn:
        paths = write_workbook(conn, selected, per_class, excluded, tmp_path)
    wb = openpyxl.load_workbook(paths["xlsx"])
    rejected_rows = list(wb["rejected_records"].iter_rows(min_row=2, values_only=True))
    quota_excluded = [r for r in rejected_rows if r[1] == "quota"]
    assert len(quota_excluded) == 3


# --- high-value columns are guaranteed + carry source_class (2026-08-12) ---


def _candidate_with_claims(entity_id="e1", claims=None):
    """Distinct from _candidate() above, which builds by score/class rather than by ledger."""
    from app.dataset import build_candidate

    entity = {"entity_id": entity_id, "canonical_name": "Acme FO"}
    return build_candidate(entity, claims or [], "ship", "SFO")


def test_high_value_columns_exist_even_when_no_record_has_the_field():
    """Columns used to be derived only from fields present in the ledger, so when none of
    the shipped records had an email the principal_email columns vanished from the sheet
    entirely — the "silently empty" outcome plan §6.3 forbids, and worse than a blank cell
    because a consumer cannot tell "looked, could not verify" from "not in this dataset".
    Observed on the real 5-record output, where 0/5 carried any contact field."""
    from app.dataset import _records_rows

    cand = _candidate_with_claims(claims=[
        {"field_name": "aum_usd", "answer": 100, "status": "confirmed",
         "source_class": "13f_filing", "confidence": "high"},
    ])
    cols, rows = _records_rows([cand])
    for f in ("principal_email", "principal_phone", "principal_name", "important_insight"):
        assert f in cols, f"{f} must always be a column"
        assert f"{f}_status" in cols
        assert f"{f}_source_class" in cols
    assert rows[0]["principal_email"] is None
    assert rows[0]["principal_email_status"] == "could_not_verify"
    assert rows[0]["principal_email_source_class"] is None


def test_source_class_sits_next_to_the_value():
    """A principal_email from `snov` is a different proposition from one scraped off
    `site_scrape`; the distinction must not require cross-referencing the provenance sheet."""
    from app.dataset import _records_rows

    cand = _candidate_with_claims(claims=[
        {"field_name": "principal_email", "answer": "jane@acme.com", "status": "confirmed",
         "source_class": "snov", "confidence": "medium"},
        {"field_name": "principal_phone", "answer": "+15551234567", "status": "format_only",
         "source_class": "site_scrape", "confidence": "low"},
    ])
    _, rows = _records_rows([cand])
    assert rows[0]["principal_email"] == "jane@acme.com"
    assert rows[0]["principal_email_source_class"] == "snov"
    assert rows[0]["principal_phone_source_class"] == "site_scrape"


def test_release_rule_blanks_the_value_but_keeps_its_source_class():
    """The release rule kills the value, not the record of where it came from — the value
    itself is preserved in audit_rejected_values. Seen live on First PREMIER Bank, whose
    Snov-sourced email was blanked as removed_failed_validation."""
    from app.dataset import _records_rows

    cand = _candidate_with_claims(claims=[
        {"field_name": "principal_email", "answer": "bad@acme.com",
         "status": "removed_failed_validation", "source_class": "snov", "confidence": "low"},
    ])
    _, rows = _records_rows([cand])
    assert rows[0]["principal_email"] is None
    assert rows[0]["principal_email_status"] == "removed_failed_validation"
    assert rows[0]["principal_email_source_class"] == "snov"


def test_non_high_value_fields_get_no_companion_columns():
    """Only high-value fields get the status/source_class pair; provenance remains the full
    per-field audit trail for everything else."""
    from app.dataset import _records_rows

    cand = _candidate_with_claims(claims=[
        {"field_name": "recent_news", "answer": "headline", "status": "confirmed",
         "source_class": "news_article", "confidence": "low"},
    ])
    cols, _ = _records_rows([cand])
    assert "recent_news" in cols
    assert "recent_news_status" not in cols
    assert "recent_news_source_class" not in cols


def test_multiple_principal_names_are_joined_not_dropped():
    """Layer D used last-write-wins per field, so a firm with two co-managing members
    shipped only the last one — silently discarding a decision-maker."""
    from app.dataset import _records_rows

    cand = _candidate_with_claims(claims=[
        {"field_name": "principal_name", "answer": "Mary C. McNutt", "status": "confirmed",
         "source_class": "web_page", "confidence": "high"},
        {"field_name": "principal_name", "answer": "Michelle J. Blass", "status": "confirmed",
         "source_class": "web_page", "confidence": "high"},
    ])
    _, rows = _records_rows([cand])
    assert rows[0]["principal_name"] == "Mary C. McNutt; Michelle J. Blass"


def test_duplicate_principal_names_are_deduped():
    from app.dataset import _records_rows

    cand = _candidate_with_claims(claims=[
        {"field_name": "principal_name", "answer": "Jane Doe", "status": "confirmed",
         "source_class": "fec_employer", "confidence": "high"},
        {"field_name": "principal_name", "answer": "Jane Doe", "status": "confirmed",
         "source_class": "web_page", "confidence": "high"},
    ])
    _, rows = _records_rows([cand])
    assert rows[0]["principal_name"] == "Jane Doe"


def test_single_valued_field_is_unaffected_by_multi_value_handling():
    from app.dataset import _records_rows

    cand = _candidate_with_claims(claims=[
        {"field_name": "aum_usd", "answer": 1, "status": "confirmed",
         "source_class": "13f_filing", "confidence": "high"},
        {"field_name": "aum_usd", "answer": 2, "status": "confirmed",
         "source_class": "13f_filing", "confidence": "high"},
    ])
    _, rows = _records_rows([cand])
    assert rows[0]["aum_usd"] == 2, "last write still wins for single-valued fields"


# --- T17.3: regression for "a column exists but no producer ever fills it" ---

# `type_final` shipped as a records-sheet column from day one but every caller had
# to invent the value (nothing persisted it), and `lead_origin_source_class` is a
# brand-new column for the same class of defect. Both must be produced from the
# claim ledger the sheet already pivots, never invented by the caller. These
# tests assert on the EMITTED SHEET ROW, not the intermediate candidate object,
# because the defect being locked out is specifically at the sheet-assembly seam.

def test_records_row_derives_type_final_and_lead_origin_from_ledger(db_path):
    from app.dataset import _records_rows, gather_survivors

    with connection(db_path) as conn:
        upsert_entity(conn, "e1", "Acme Capital")
        upsert_claims(conn, "e1", [
            _claim(None, question_id="G1.Q4", status="single_source", answer="It is a multi-family office (MFO)."),
            _claim("discovery_class_13f_filing", answer=True, status="confirmed"),
        ])
        # 2-tuple: NO caller-supplied type_final — it must be derived from the ledger.
        candidates = gather_survivors(conn, [("e1", "ship")])
    cols, rows = _records_rows(candidates)
    row = rows[0]
    assert row["type_final"] == "MFO"
    assert row["lead_origin_source_class"] == "13f_filing"
    # the column sits immediately after type_final in the emitted header
    assert "lead_origin_source_class" in cols
    assert cols.index("lead_origin_source_class") == cols.index("type_final") + 1


def test_records_row_unknown_class_and_unconfirmed_type_when_ledger_lacks_them(db_path):
    from app.dataset import _records_rows, gather_survivors

    with connection(db_path) as conn:
        upsert_entity(conn, "e2", "Mystery Capital")
        upsert_claims(conn, "e2", [
            _claim(None, question_id="G1.Q4", status="could_not_verify", answer=None),
            # no discovery_class_* claim at all
        ])
        candidates = gather_survivors(conn, [("e2", "ship")])
    _, rows = _records_rows(candidates)
    assert rows[0]["type_final"] == "type_unconfirmed"
    assert rows[0]["lead_origin_source_class"] == "unknown"
