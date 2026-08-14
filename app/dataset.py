"""Layer D — Dataset assembly (enrichment_validation_dataset_plan.md §6). Selection
(quota.py's logic, §6.1) -> six-sheet assembly (§6.2) -> XLSX/CSV/manifest emission
(§6.3). Everything reads from the `claims` spine (app/db.py) — "the records sheet is
literally a pivot of the ledger" (plan §6.2).

No scoring formula for `actionability_score` / `urgency_tier_rank` is specified
anywhere in this repo (the plan's §6.1 pseudocode references them as already-computed
fields on a `survivors` object that doesn't exist elsewhere) — `_score_claims` and
`_urgency_tier_rank` below are a documented, defensible proxy: see their docstrings.
"""
from __future__ import annotations

import csv
import json
import logging
import os
import subprocess
from collections import Counter
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook

from app.db import (
    get_audit_rejected_values,
    get_claims,
    get_entity,
    get_findings,
    get_rejections,
    write_production_record,
    start_run,
    finish_run,
    write_field_provenance,
)
from app.validation import _MULTI_VALUED_FIELDS, decide_type_final

logger = logging.getLogger(__name__)

MAX_PER_CLASS = 15  # 30% of 50 — plan §6.1

_STATUS_WEIGHTS: dict[str, int] = {
    "verified": 3, "confirmed": 2, "single_source": 2, "format_only": 1, "pattern_inferred": 1,
}

# Column order: identity -> type -> principal contact -> why-now -> mandates -> signals
# -> integrity (plan §6.3). Any field present in the ledger but not listed here is
# appended at the end, never dropped.
_COLUMN_ORDER: list[str] = [
    # identity
    "entity_id", "canonical_name", "domain",
    # type
    "type_final", "aum_usd", "aum_basis", "aum_as_of", "headcount",
    # principal contact
    "principal_name", "principal_title", "principal_email", "principal_phone",
    "principal_linkedin", "corporate_linkedin",
    "secondary_contact_name", "secondary_contact_email", "secondary_contact_phone",
    # why-now
    "important_insight", "recent_investments", "recent_fund_commitments",
    "recent_key_hires", "recent_news",
    # mandates
    "investing_mandates", "investing_thesis", "sector_focus", "stage_focus",
    "geography_focus", "check_size_range", "direct_vs_fund", "do_not_pitch",
    "fit_tags", "background",
    # integrity
    "public_list_overlap",
]

_HIGH_VALUE_FIELDS: frozenset[str] = frozenset({
    "principal_name", "principal_email", "principal_phone", "aum_usd", "important_insight",
})

# T26 — provenance tiers for multi-valued fields. An extraction_method PREFIX maps to a
# rank; lower is better. `_provenance_rank` does longest-prefix match. The point: a
# researcher-verified `projected_G2.Q1` principal must not be concatenated in the same
# cell as `derived_fec_employer` donors (campaign contributors who named the firm as their
# employer). `projected_` is a Layer-1 researched, source-cited finding; `derived_` is a raw
# discovery-feed payload value; everything else (serper_xray, serper_organic, jsonld,
# site_scrape, and anything unrecognised) is a search-snippet / scrape tier. `None` ranks
# 2 — unknown provenance is not evidence of quality.
_PROVENANCE_TIER: dict[str, int] = {
    "projected_": 0,
    "derived_": 1,
}


def _provenance_rank(extraction_method: str | None) -> int:
    """Longest-prefix match against `_PROVENANCE_TIER`; unmatched and `None` -> 2."""
    if extraction_method is None:
        return 2
    best = 2
    best_prefix_len = -1
    for prefix, rank in _PROVENANCE_TIER.items():
        if extraction_method.startswith(prefix) and len(prefix) > best_prefix_len:
            best = rank
            best_prefix_len = len(prefix)
    return best


@dataclass
class ProductionCandidate:
    entity_id: str
    canonical_name: str
    type_final: str
    outcome: str
    claims: list[dict[str, Any]] = field(default_factory=list)
    actionability_score: float = 0.0
    verified_cell_count: int = 0
    urgency_tier_rank: int = 0
    discovery_class_primary: str | None = None
    discovery_class_count: int = 0


def _score_claims(claims: list[dict[str, Any]]) -> tuple[float, int]:
    """actionability_score + verified_cell_count. Sum of per-status weights across
    every field-bearing (field_name is not None) settled claim — "verified" (cross-
    class corroborated) outweighs a merely "confirmed"/"single_source" claim, since
    that's the entire point of the cross-class rule (plan §5): a record with more
    independently-corroborated cells is more actionable than one with the same cell
    count all from one source."""
    score = 0.0
    verified = 0
    for c in claims:
        if not c.get("field_name"):
            continue
        score += _STATUS_WEIGHTS.get(c["status"], 0)
        if c["status"] == "verified":
            verified += 1
    return score, verified


def _urgency_tier_rank(claims: list[dict[str, Any]]) -> int:
    """0 = no dated signal at all, 1 = a low-confidence dated signal only, 2 = a
    medium/high-confidence dated signal. Used purely as a selection tiebreaker."""
    signal_fields = {"important_insight", "recent_investments", "recent_fund_commitments"}
    signals = [
        c for c in claims
        if c.get("field_name") in signal_fields and c["status"] not in ("could_not_verify", "removed_failed_validation")
    ]
    if not signals:
        return 0
    return 2 if any(c.get("confidence") in ("high", "medium") for c in signals) else 1


def _discovery_class_info(claims: list[dict[str, Any]]) -> tuple[str | None, int]:
    classes = sorted(
        c["field_name"][len("discovery_class_"):]
        for c in claims
        if (c.get("field_name") or "").startswith("discovery_class_")
    )
    return (classes[0] if classes else None, len(classes))


def build_candidate(entity: dict[str, Any], claims: list[dict[str, Any]], outcome: str, type_final: str | None = None) -> ProductionCandidate:
    # type_final defaults to None = "derive it from the ledger", distinct from a
    # literal "type_unconfirmed" = "we decided it is unconfirmed". Layer V (which
    # has already run the rule and persisted nothing) can still pass an explicit
    # value and have it win — that path is what the existing tests exercise.
    if type_final is None:
        type_final = decide_type_final(claims)
    score, verified = _score_claims(claims)
    primary, count = _discovery_class_info(claims)
    return ProductionCandidate(
        entity_id=entity["entity_id"], canonical_name=entity["canonical_name"],
        type_final=type_final, outcome=outcome, claims=claims,
        actionability_score=score, verified_cell_count=verified,
        urgency_tier_rank=_urgency_tier_rank(claims),
        discovery_class_primary=primary, discovery_class_count=count,
    )


def select_50(
    survivors: list[ProductionCandidate], n: int = 50, max_per_class: int = MAX_PER_CLASS
) -> tuple[list[ProductionCandidate], dict[str, int], list[ProductionCandidate]]:
    """§6.1 — ranked, then quota-bound. Returns (selected, per_class_counts,
    excluded_by_quota). `excluded_by_quota` is returned (not just logged) so the caller
    can persist it to `production_records` — "log it — 'excluded to preserve source
    diversity' is a defensible, documented judgment call" (plan §6.1)."""
    ranked = sorted(
        survivors,
        key=lambda r: (
            r.actionability_score,
            r.type_final == "SFO",
            r.verified_cell_count,
            r.urgency_tier_rank,
            r.discovery_class_count,
        ),
        reverse=True,
    )
    selected: list[ProductionCandidate] = []
    excluded: list[ProductionCandidate] = []
    per_class: Counter[str] = Counter()
    for r in ranked:
        if len(selected) >= n:
            break
        cls = r.discovery_class_primary or "unknown"
        if per_class[cls] >= max_per_class:
            excluded.append(r)
            continue
        selected.append(r)
        per_class[cls] += 1
    return selected, dict(per_class), excluded


def persist_selection(
    conn, selected: list[ProductionCandidate], excluded: list[ProductionCandidate]
) -> None:
    for rank, r in enumerate(selected, start=1):
        write_production_record(conn, r.entity_id, rank=rank, primary_class=r.discovery_class_primary,
                                 excluded_by_quota=False)
    for r in excluded:
        write_production_record(conn, r.entity_id, rank=None, primary_class=r.discovery_class_primary,
                                 excluded_by_quota=True)


def gather_survivors(conn, entity_outcomes: list[tuple[str, str] | tuple[str, str, str]]) -> list[ProductionCandidate]:
    """`entity_outcomes` is [(entity_id, outcome[, type_final]), ...] for ship/
    ship_with_caveats entities — the caller (app.enrichment.run_pipeline's summary,
    or a fresh DB scan) is the source of truth for which entities reached which
    outcome, since that isn't itself persisted as a queryable column anywhere. A
    2-tuple (entity_id, outcome) means "derive type_final from the ledger"; a
    3-tuple supplies it explicitly and wins. Either way `type_final` is now produced
    from the claim ledger, never invented by the caller — the column used to ship
    empty because no persisted value existed and every caller had to guess."""
    candidates = []
    for item in entity_outcomes:
        entity_id, outcome = item[0], item[1]
        type_final = item[2] if len(item) > 2 else None
        entity = get_entity(conn, entity_id)
        if entity is None:
            continue
        claims = get_claims(conn, entity_id)
        candidates.append(build_candidate(entity, claims, outcome, type_final))
    return candidates


# ============================================================================
# 6.2 — sheet assembly
# ============================================================================


@dataclass
class CellResolution:
    """The decision for one cell, extracted so the row (`_records_rows`) and the provenance
    log (`app.provenance_log.build_field_records`) can never disagree on what a cell
    contains — the T19 vocabulary split happened because two code paths each derived a
    field's value and drifted. One function, two callers (PLAN.md T35.3).

    - `value` is exactly what the row cell gets (None = blank).
    - `winner` is the claim whose `status`/`source_class` the row's companion columns
      carry (last write wins, same convention as the pre-refactor `_records_rows`). For a
      multi-valued field the value can come from entirely different claims than the
      winner — `winner` is the row's status representative, NOT necessarily the value's
      source. The provenance log attributes `how`/`verification`/`confidence`/`status` to
      `producers[0]`, not to `winner`, so the log never misattributes a shipped value to a
      claim that did not produce it (the single failure mode this feature exists to
      prevent).
    - `producers` are the claims whose values are actually in the cell, in cell order:
      `[winner]` for a single-valued field with a non-null value (empty for a blanked /
      claim-less cell); the best-tier claims that contributed a de-duplicated answer for a
      multi-valued field. Empty when the cell is blank.
    - `alternatives` pairs every field claim that is NOT in `producers` with a
      `why_not_used` code from the closed vocabulary (lower_provenance_tier | superseded |
      failed_validation | not_latest_write | duplicate_value | weaker_status) — including
      the winner when the winner did not produce the value.
    - `multi_values` is the joined parts for a multi-valued field, else None.
    """
    value: Any
    winner: dict[str, Any] | None
    alternatives: list[tuple[dict[str, Any], str]]
    multi_values: list[str] | None
    producers: list[dict[str, Any]]


def resolve_cell(claims: list[dict[str, Any]], field: str) -> CellResolution:
    """Decide what one cell contains, purely from that entity's claim list. Reproduces
    the pre-refactor `_records_rows` rules EXACTLY for `value`/`winner`/`multi_values` —
    do not "improve" a rule here; the acceptance test is that the emitted sheet does not
    change by one byte.

    single-valued: last write wins; a winner whose status is `removed_failed_validation`
    yields value=None (the value was killed, the claim is still the winner and is reported
    as such so the companion `_status` column keeps `removed_failed_validation`).
    multi-valued (`_MULTI_VALUED_FIELDS`): bucket eligible claims by
    `_provenance_rank(extraction_method)`, keep ONLY the lowest (best) rank present,
    first-seen order within that rank, de-duplicated, "; "-joined. Claims with status
    `removed_failed_validation` or `superseded` are excluded from the buckets, and an
    empty/None answer is skipped — same as today.

    `producers` is new (T35.3 D1): the claims whose values are in the cell, so the
    provenance log can attribute the shipped value to the claim that actually produced it
    rather than to the last-written claim (which for a multi-valued field may be a
    worse-tier claim that contributed nothing).
    """
    field_claims = [c for c in claims if c.get("field_name") == field]
    if not field_claims:
        return CellResolution(value=None, winner=None, alternatives=[],
                              multi_values=None, producers=[])

    # Last write wins — the last claim in iteration order with this field_name. This is
    # the claim whose status/source_class the row's companion columns carry, for both
    # single- and multi-valued fields (the pre-refactor code used `latest[f]` for the
    # companion columns regardless of multi-valued). Must stay last-write to keep the
    # emitted sheet byte-identical. NOTE: for a multi-valued field the winner may NOT be
    # one of the producers — the row's _status column reflects the last write, the value
    # reflects the best provenance tier, and those can be different claims.
    winner = field_claims[-1]
    multi_values: list[str] | None = None
    producers: list[dict[str, Any]] = []
    # rank of each eligible claim (status ok, non-empty answer), so the alternatives
    # coding can distinguish a worse-tier claim from a duplicate answer in the winning
    # tier (D2: two independent sources agreeing is corroboration, not a rejection).
    eligible_rank: dict[int, int] = {}
    tiers: dict[int, list[str]] = {}
    best: int | None = None

    if field in _MULTI_VALUED_FIELDS:
        # {rank: [answers in first-seen order]}. T26: build the cell from ONLY the best
        # (lowest) rank present, so a researcher-verified `projected_G2.Q1` principal is
        # not concatenated in the same cell as `derived_fec_employer` donors.
        tier_claims: dict[int, list[dict[str, Any]]] = {}
        for cl in field_claims:
            if cl.get("status") in ("removed_failed_validation", "superseded"):
                continue
            answer = cl.get("answer")
            if answer in (None, ""):
                continue
            rank = _provenance_rank(cl.get("extraction_method"))
            eligible_rank[id(cl)] = rank
            bucket = tiers.setdefault(rank, [])
            if str(answer) not in bucket:
                bucket.append(str(answer))
                tier_claims.setdefault(rank, []).append(cl)
        if tiers:
            best = min(tiers)
            parts = tiers[best]
            multi_values = list(parts)
            value: Any = "; ".join(parts)
            # the best-tier claims that contributed a de-duplicated answer, in cell order.
            producers = list(tier_claims[best])
        else:
            # No eligible multi claim -> falls back to the single-valued rule using the
            # winner (last write), matching the pre-refactor `else` branch.
            value = winner.get("answer") if winner.get("status") != "removed_failed_validation" else None
            producers = [winner] if value is not None else []
    else:
        value = winner.get("answer") if winner.get("status") != "removed_failed_validation" else None
        producers = [winner] if value is not None else []

    producing_set = {id(c) for c in producers}

    # alternatives: every field claim NOT in `producers` — including the winner when the
    # winner did not produce the value (D1). A last-write claim that contributed nothing
    # is a rejected alternative, not a silently-dropped one.
    alternatives: list[tuple[dict[str, Any], str]] = []
    for cl in field_claims:
        if id(cl) in producing_set:
            continue
        # An alternative is a VALUE that lost. A claim carrying no answer at all lost
        # nothing — it is the record of a lookup that came back empty, and the cell's
        # `blank_reason` already tells that story properly. Listing it here rendered as
        # "rejected: null — not_latest_write" in the Log tab: noise that reads like a
        # finding. Seen live on a `principal_phone` whose only claim was a
        # tool_unavailable could_not_verify.
        if cl.get("answer") in (None, ""):
            continue
        status = cl.get("status")
        if status == "superseded":
            code = "superseded"
        elif status == "removed_failed_validation":
            code = "failed_validation"
        elif field in _MULTI_VALUED_FIELDS and tiers:
            rank = eligible_rank.get(id(cl))
            if rank is not None and rank == best:
                # eligible, in the winning tier, but its answer was already shipped from
                # an earlier claim in the same tier — two independent sources agreeing is
                # corroboration (D2), not a rejection.
                code = "duplicate_value"
            elif rank is not None and rank > best:
                # a multi-valued claim in a worse provenance bucket than the one that shipped.
                code = "lower_provenance_tier"
            else:
                # an empty-answer claim or anything else that lost without a clearer code.
                code = "weaker_status"
        else:
            # single-valued claim beaten by a later write for the same field.
            code = "not_latest_write"
        alternatives.append((cl, code))

    # Deterministic order (created_at, then claim_id) — a web page renders this list and
    # it must not reshuffle between reads.
    alternatives.sort(key=lambda pair: (pair[0].get("created_at") or "", pair[0].get("claim_id") or ""))

    return CellResolution(value=value, winner=winner, alternatives=alternatives,
                          multi_values=multi_values, producers=producers)


def _records_rows(selected: list[ProductionCandidate]) -> tuple[list[str], list[dict[str, Any]]]:
    """One row per entity, one column per field_name, plus `<field>_status` and
    `<field>_source_class` companion columns for every high-value field — plan §6.3:
    "populated, or blank with could_not_verify status column. Never silently empty."

    Two things this function gets deliberately right, both learned the hard way:

    1. **Every high-value field ALWAYS gets its columns**, even when no selected record has
       a claim for it. Columns used to be derived purely from fields present in the ledger,
       so when none of the shipped records had an email the `principal_email` and
       `principal_email_status` columns vanished from the sheet entirely. That is exactly
       the "silently empty" outcome the plan forbids — and it is worse than a blank cell,
       because a consumer cannot distinguish "we looked and could not verify it" from "this
       field is not part of the dataset". Observed on the current 5-record output, where
       0/5 records carried any contact field.

    2. **`<field>_source_class` sits next to the value**, so a reader can judge a cell
       without cross-referencing the `provenance` sheet. A `principal_email` sourced from
       `snov` is a different proposition from one scraped off `site_scrape`, and the whole
       point of the claim ledger is that the distinction stays attached to the value.
       (`provenance` remains the full audit trail — every field, plus confirming_url,
       confirming_class and verification_method.)

    3. **The companion columns describe the claim that PRODUCED the value** (T35.7), not
       whichever claim happened to be written last. They used to read `res.winner`, which
       is the last write — correct for a single-valued field, wrong for a multi-valued one
       whose cell is filled from the best provenance tier. Measured case: a row shipping
       `principal_name = "Matt Blackburn"` (a researcher's `projected_G2.Q1` site-scrape
       claim, which wins on tier) carried `principal_name_source_class = "fec_employer"` —
       the class of a campaign-donor claim for a different person that merely arrived
       later. The value was right and the label was wrong, in the deliverable itself.
       `res.producers[0]` is the claim whose value is actually in the cell; `res.winner`
       remains the fallback for a blanked cell (a `removed_failed_validation` winner
       produces nothing, and its status is exactly what the reader needs to see).
    """
    present_fields: set[str] = set()
    for c in selected:
        present_fields.update(cl["field_name"] for cl in c.claims if cl.get("field_name"))

    # High-value fields are guaranteed columns whether or not anyone has one.
    all_fields = present_fields | set(_HIGH_VALUE_FIELDS)
    ordered = [f for f in _COLUMN_ORDER if f in all_fields]
    ordered += sorted(all_fields - set(ordered))

    columns: list[str] = ["entity_id", "canonical_name", "type_final", "lead_origin_source_class", "outcome"]
    for f in ordered:
        if f in ("entity_id", "canonical_name", "type_final", "lead_origin_source_class"):
            continue
        columns.append(f)
        if f in _HIGH_VALUE_FIELDS:
            columns.append(f"{f}_status")
            columns.append(f"{f}_source_class")

    rows = []
    for c in selected:
        # lead_origin_source_class is a candidate-level attribute (which discovery feed
        # this lead came from), NOT a claim field_name — so it follows the type_final
        # precedent: a literal header entry + a literal row entry, never derived from
        # the claim field_name pivot. Same value that drives the select_50 per-class
        # quota (c.discovery_class_primary); 'unknown' when no discovery_class_* claim.
        row: dict[str, Any] = {"entity_id": c.entity_id, "canonical_name": c.canonical_name,
                                "type_final": c.type_final,
                                "lead_origin_source_class": c.discovery_class_primary or "unknown",
                                "outcome": c.outcome}
        for f in ordered:
            if f in ("entity_id", "canonical_name", "type_final", "lead_origin_source_class"):
                continue
            # One decision, two consumers: the cell value comes from resolve_cell, the
            # same function app.provenance_log.build_field_records calls, so the row and
            # the provenance log can never disagree on what a cell contains (T35.3 — the
            # no-drift move T27.2 made by importing _provenance_rank rather than
            # re-implementing it).
            res = resolve_cell(c.claims, f)
            row[f] = res.value
            if f in _HIGH_VALUE_FIELDS:
                # T35.7: describe the claim that PRODUCED the shipped value. `producers[0]`
                # is that claim; it differs from `winner` (the last write) exactly when a
                # multi-valued cell was filled from a better provenance tier than the
                # last-written claim — the case that used to label a researcher-sourced
                # principal `fec_employer`. A blanked cell has no producer, so it falls
                # back to the winner, whose status is the very thing the reader needs
                # (`removed_failed_validation`, `could_not_verify`, …).
                describing = res.producers[0] if res.producers else res.winner
                row[f"{f}_status"] = describing["status"] if describing else "could_not_verify"
                # A blanked value keeps its source_class: the release rule killed the value,
                # not the record of where it came from (the value itself is preserved in
                # audit_rejected_values).
                row[f"{f}_source_class"] = describing.get("source_class") if describing else None
        rows.append(row)
    return columns, rows


def _provenance_rows(selected: list[ProductionCandidate]) -> list[dict[str, Any]]:
    """One row per (record, field) — unpivoted, full provenance."""
    rows = []
    for c in selected:
        for cl in c.claims:
            if not cl.get("field_name"):
                continue
            rows.append({
                "entity_id": c.entity_id,
                "field_name": cl["field_name"],
                "answer": cl["answer"],
                "status": cl["status"],
                "source_url": cl.get("source_url"),
                "source_class": cl.get("source_class"),
                "extraction_method": cl.get("extraction_method"),
                "confidence": cl.get("confidence"),
                "verification_method": cl.get("verification_method"),
                "confirming_url": cl.get("confirming_url"),
                "confirming_class": cl.get("confirming_class"),
                "verified_at": cl.get("verified_at"),
            })
    return rows


def _source_class_report_rows(selected: list[ProductionCandidate]) -> list[dict[str, Any]]:
    per_class: dict[str, dict[str, Any]] = {}
    for c in selected:
        cls = c.discovery_class_primary or "unknown"
        bucket = per_class.setdefault(cls, {"discovery_class": cls, "entity_count": 0, "score_sum": 0.0})
        bucket["entity_count"] += 1
        bucket["score_sum"] += c.actionability_score
    rows = []
    for cls, bucket in sorted(per_class.items()):
        rows.append({
            "discovery_class": cls,
            "entity_count": bucket["entity_count"],
            "avg_actionability_score": round(bucket["score_sum"] / bucket["entity_count"], 2) if bucket["entity_count"] else 0,
        })
    return rows


_DATA_DICTIONARY: list[dict[str, str]] = [
    {"field": "aum_usd", "description": "Assets under management, USD floor derived from 13F tableValueTotal.", "inclusion_standard": "derived (wave -1); never fabricated if no 13F filing exists."},
    {"field": "aum_basis", "description": "How aum_usd was computed — currently always '13f_floor' (a floor, not a total AUM figure).", "inclusion_standard": "always paired with aum_usd."},
    {"field": "principal_name", "description": "Named decision-maker (principal, CIO, or similar).", "inclusion_standard": "wave 1 actionability-core field; a record without one is gated before wave 2 spend."},
    {"field": "principal_email", "description": "Contact email for the principal or the firm.", "inclusion_standard": "release rule: an email that fails verification is blanked, status='removed_failed_validation', logged to audit_rejected_values."},
    {"field": "principal_phone", "description": "Contact phone.", "inclusion_standard": "format-valid only unless independently corroborated — see status='format_only' vs 'verified'."},
    {"field": "important_insight", "description": "The reason this is a live opportunity now (concentration_pain | fresh_liquidity | access_window).", "inclusion_standard": "derived from 13F QoQ deltas or a future-dated conference sighting; never guessed."},
    {"field": "type_final", "description": "SFO | MFO | type_unconfirmed, from the G1.Q4 claim's final validated status.", "inclusion_standard": "type_unconfirmed if G1.Q4 was never settled or was contradicted."},
    {"field": "lead_origin_source_class", "description": "Discovery feed class this lead originally came from (13f_filing | fec_employer | 5500_filing | ppp_loans); 'unknown' if the lead carries no discovery_class_* claim.", "inclusion_standard": "derived from the discovery_class_* claims written by wave -1; the same value that drives the select_50 per-class quota."},
]

_STATUS_VOCAB_NOTES: list[dict[str, str]] = [
    {"status": "verified", "meaning": "Confirmed by >=2 independent, cross-class sources (a search-index hit alone never counts)."},
    {"status": "single_source", "meaning": "Confirmed, but only ever seen from one source_class — not independently corroborated."},
    {"status": "pattern_inferred", "meaning": "Format-plausible but unconfirmed (e.g. a pattern-guessed email)."},
    {"status": "format_only", "meaning": "Passes a format check only (e.g. phone number shape) — NOT line-verified."},
    {"status": "could_not_verify", "meaning": "Genuinely absent, or the tool that would have found it was unavailable — see extraction_method."},
    {"status": "removed_failed_validation", "meaning": "Killed by the release rule; original value is in audit_rejected_values, not here."},
]


# ============================================================================
# 6.3 — emission
# ============================================================================


def _git_sha() -> str:
    try:
        return subprocess.check_output(
            ["git", "rev-parse", "HEAD"], cwd=Path(__file__).resolve().parent.parent, timeout=5,
        ).decode().strip()
    except Exception:
        return "unknown"


def _atomic_write_json(path: Path, obj: Any) -> None:
    """Write JSON atomically: a temp file in the same directory, then ``os.replace``. A
    half-written file must never be served — a web page reading
    ``field_provenance.json`` mid-write would otherwise see truncated JSON. UTF-8 with
    ``ensure_ascii=False`` because scraped content carries non-ASCII punctuation."""
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_name(path.name + ".tmp")
    tmp.write_text(json.dumps(obj, indent=2, ensure_ascii=False, default=str), encoding="utf-8")
    os.replace(tmp, path)


def write_workbook(
    conn,
    selected: list[ProductionCandidate],
    per_class: dict[str, int],
    excluded: list[ProductionCandidate],
    output_dir: str | Path,
    *,
    budget_spent: dict[str, Any] | None = None,
    run_id: str | None = None,
) -> dict[str, str]:
    """Writes the six-sheet XLSX, records.csv, and run_manifest.json to `output_dir`.
    Returns {"xlsx": path, "csv": path, "manifest": path, "field_provenance": path}.

    T35.5: after the sheets are written, builds and persists the per-run field
    provenance log for ``selected + excluded`` and writes ``field_provenance.json``.
    ``run_id`` owns the log; when None a ``dataset`` run is opened (and closed) here so
    the log always has an owning run. An emission failure is logged and does NOT fail
    the workbook write — the leads are the deliverable, the log is the record of it."""
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    columns, record_rows = _records_rows(selected)
    provenance_rows = _provenance_rows(selected)
    audit_rows = get_audit_rejected_values(conn)
    rejected_rows = get_rejections(conn)
    source_class_rows = _source_class_report_rows(selected)

    wb = Workbook()
    ws = wb.active
    ws.title = "records"
    ws.append(columns)
    for row in record_rows:
        ws.append([row.get(col) for col in columns])

    ws2 = wb.create_sheet("provenance")
    prov_cols = ["entity_id", "field_name", "answer", "status", "source_url", "source_class",
                 "extraction_method", "confidence", "verification_method", "confirming_url",
                 "confirming_class", "verified_at"]
    ws2.append(prov_cols)
    for row in provenance_rows:
        ws2.append([_stringify(row.get(c)) for c in prov_cols])

    ws3 = wb.create_sheet("audit_rejected_values")
    audit_cols = ["entity_id", "field_name", "rejected_value", "reason_code", "evidence_url", "rejected_at"]
    ws3.append(audit_cols)
    for row in audit_rows:
        ws3.append([_stringify(row.get(c)) for c in audit_cols])

    ws4 = wb.create_sheet("rejected_records")
    rej_cols = ["entity_id", "stage", "reason_code", "created_at"]
    ws4.append(rej_cols)
    for row in rejected_rows:
        ws4.append([row.get(c) for c in rej_cols])
    for r in excluded:
        ws4.append([r.entity_id, "quota", "excluded to preserve source diversity", None])

    ws5 = wb.create_sheet("source_class_report")
    sc_cols = ["discovery_class", "entity_count", "avg_actionability_score"]
    ws5.append(sc_cols)
    for row in source_class_rows:
        ws5.append([row.get(c) for c in sc_cols])
    ws5.append([])
    ws5.append(["quota cap per class", MAX_PER_CLASS])
    for cls, count in sorted(per_class.items()):
        ws5.append([f"selected: {cls}", count])

    ws6 = wb.create_sheet("data_dictionary")
    ws6.append(["field", "description", "inclusion_standard"])
    for row in _DATA_DICTIONARY:
        ws6.append([row["field"], row["description"], row["inclusion_standard"]])
    ws6.append([])
    ws6.append(["status", "meaning", ""])
    for row in _STATUS_VOCAB_NOTES:
        ws6.append([row["status"], row["meaning"], ""])

    xlsx_path = output_dir / "production_dataset.xlsx"
    wb.save(xlsx_path)

    csv_path = output_dir / "records.csv"
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=columns)
        writer.writeheader()
        for row in record_rows:
            writer.writerow({k: _stringify(v) for k, v in row.items()})

    manifest = {
        "git_sha": _git_sha(),
        "run_timestamp": datetime.now(timezone.utc).isoformat(),
        "selected_count": len(selected),
        "excluded_by_quota_count": len(excluded),
        "per_class_counts": per_class,
        "budget_spent": budget_spent or {},
    }

    # T35.5: build + persist the per-run field provenance log for selected + excluded
    # and write field_provenance.json. Lazy import avoids a cycle (app.provenance_log
    # imports from app.dataset). An emission failure must never fail the workbook write —
    # the leads are the deliverable, the log is the record of it.
    provenance_path: Path | None = None
    owns_run = run_id is None
    if owns_run:
        run_id = start_run(conn, "dataset", entity_count=len(selected) + len(excluded))
    try:
        from app.provenance_log import build_run_log
        entity_outcomes = [(c.entity_id, c.outcome) for c in selected + excluded]
        doc = build_run_log(conn, run_id, entity_outcomes)
        rows = []
        for lead in doc["leads"]:
            for rec in lead["fields"]:
                rows.append({
                    "run_id": run_id,
                    "entity_id": lead["entity_id"],
                    "field": rec["field"],
                    "value": rec["value"],
                    "status": rec["status"],
                    "shipped": rec["shipped"],
                    "source_class": (rec["how"] or {}).get("source_class"),
                    "extraction_method": (rec["how"] or {}).get("extraction_method"),
                    "record": json.dumps(rec, ensure_ascii=False, default=str),
                })
        write_field_provenance(conn, rows)
        provenance_path = output_dir / "field_provenance.json"
        _atomic_write_json(provenance_path, doc)
        manifest["field_provenance"] = str(provenance_path)
        if owns_run:
            finish_run(conn, run_id, status="done",
                       notes={"selected": len(selected), "excluded": len(excluded)})
    except Exception:  # noqa: BLE001 — never fail the run over the log
        logger.error("field provenance emission failed for run_id=%s", run_id, exc_info=True)
        if owns_run:
            try:
                finish_run(conn, run_id, status="failed")
            except Exception:  # noqa: BLE001
                pass

    manifest_path = output_dir / "run_manifest.json"
    manifest_path.write_text(json.dumps(manifest, indent=2, default=str), encoding="utf-8")

    out = {"xlsx": str(xlsx_path), "csv": str(csv_path), "manifest": str(manifest_path)}
    if provenance_path is not None:
        out["field_provenance"] = str(provenance_path)
    return out


def _stringify(value: Any) -> Any:
    if isinstance(value, (dict, list)):
        return json.dumps(value, default=str)
    return value
